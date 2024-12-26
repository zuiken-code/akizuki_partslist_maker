from df_maker import df_maker
import openpyxl

def main():
    sheet_name = '購入品リスト1'
    numbers , pieces = read_xl(sheet_name)
    df = df_maker(numbers,pieces)
    print(df)
    write_xl(sheet_name,numbers,df)

def write_xl(sheet_name,numbers,df):
    workbook = openpyxl.load_workbook("購入品リスト.xlsx")
    sheet = workbook[sheet_name]
    shape = df.shape
    for i in range(shape[0]):
        index = numbers[i]
        result = df.loc[index,:]
        result = list(result)
        print(result)
        for x in range(len(result)):
            val = result[x]
            row = i + 3
            colume = x + 2
            sheet.cell(row, colume).value = val
    workbook.save("購入品リスト.xlsx")

def read_xl(sheet_name):
    numbers = []
    pieces = []
    workbook = openpyxl.load_workbook("購入品リスト.xlsx")
    sheet = workbook[sheet_name]

    i = 3 #販売コードが入り始める行番号
    while True:
        cell = sheet.cell(i, 1).value
        if cell != None:
            numbers.append(cell)
            i += 1
        else:
            break
    
    i = 3 #個数が入り始める行番号
    while True:
        cell = sheet.cell(i, 4).value
        if cell != None:
            pieces.append(cell)
            i += 1
        else:
            break
    return numbers, pieces



if __name__ == '__main__':
    main()